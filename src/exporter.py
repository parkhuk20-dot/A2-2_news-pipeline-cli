"""데이터 내보내기 — CSV / Excel(xlsx).

필터(--status summarized, --category, --date-from/to)를 적용해 원하는 부분만 뽑는다.
CSV 는 Excel 에서 한글이 깨지지 않도록 UTF-8 BOM(utf-8-sig) 으로 저장한다.
"""

from __future__ import annotations

import argparse
import csv
from datetime import datetime
from pathlib import Path

from .config import Config
from .db import Database
from .logger import get_logger

log = get_logger("export")

COLUMNS = [
    ("id", "ID"),
    ("source", "언론사"),
    ("category", "카테고리"),
    ("published_at", "발행일"),
    ("title", "제목"),
    ("url", "URL"),
    ("summary", "요약"),
    ("sentiment", "감성"),
    ("orig_len", "원문길이"),
    ("summary_len", "요약길이"),
    ("is_duplicate", "유사중복"),
    ("body", "본문"),
]

EXCEL_CELL_LIMIT = 32000


def _rows_to_records(rows) -> list[dict]:
    records = []
    for row in rows:
        record = {}
        for key, _ in COLUMNS:
            value = row[key] if key in row.keys() else None
            if isinstance(value, str) and len(value) > EXCEL_CELL_LIMIT:
                value = value[:EXCEL_CELL_LIMIT] + "…(생략)"
            record[key] = value
        records.append(record)
    return records


def export_csv(records: list[dict], path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([label for _, label in COLUMNS])
        for record in records:
            writer.writerow([record[key] for key, _ in COLUMNS])


def export_xlsx(records: list[dict], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "뉴스"

    sheet.append([label for _, label in COLUMNS])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for record in records:
        sheet.append([record[key] for key, _ in COLUMNS])

    widths = {"title": 45, "url": 40, "summary": 50, "body": 60, "published_at": 12}
    for index, (key, _) in enumerate(COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(key, 12)

    sheet.freeze_panes = "A2"
    workbook.save(path)


def run_export(args: argparse.Namespace, cfg: Config) -> int:
    with Database(cfg.path_for("db")) as db:
        rows = db.query_articles(
            category=args.category,
            date_from=args.date_from,
            date_to=args.date_to,
            status=args.status,
        )
        records = _rows_to_records(rows)

    if not records:
        log.warning("조건에 맞는 데이터가 없어 내보낼 것이 없습니다.")
        return 0

    if args.output:
        path = Path(args.output)
    else:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = f"_{args.status}" if args.status else ""
        path = cfg.path_for("exports") / f"news_{stamp}{suffix}.{args.format}"
    path.parent.mkdir(parents=True, exist_ok=True)

    filters = []
    if args.status:
        filters.append(f"status={args.status}")
    if args.category:
        filters.append(f"category={args.category}")
    if args.date_from or args.date_to:
        filters.append(f"기간={args.date_from or '처음'}~{args.date_to or '현재'}")
    log.info("내보내기 대상: %d건%s", len(records), f" ({', '.join(filters)})" if filters else "")

    if args.format == "csv":
        export_csv(records, path)
    else:
        export_xlsx(records, path)

    log.info("내보내기 완료: %s (%s, %d건)", path, args.format.upper(), len(records))
    return 0
